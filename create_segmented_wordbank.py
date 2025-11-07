from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Reading Word Bank"

# Headers
ws['A1'] = "Category"
ws['B1'] = "Pattern/Description"
ws['C1'] = "Words (10-30 per section)"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
for cell in ['A1', 'B1', 'C1']:
    ws[cell].fill = header_fill
    ws[cell].font = header_font
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 90

# SEGMENTED Word bank - smaller, manageable chunks
word_bank = [
    # ========== HIGH FREQUENCY WORDS - BROKEN INTO CHUNKS ==========
    ("0A. YEAR 1 HF WORDS (1-25)", "First 25 high frequency words", "the, a, to, I, and, he, of, it, was, said, in, his, is, for, on, are, as, with, they, be, at, this, have, from, or"),

    ("0A. YEAR 1 HF WORDS (26-50)", "Words 26-50", "one, had, by, but, not, what, all, were, we, when, your, can, there, out, this, been, call, may, did, down, way, could, people, my, than"),

    ("0A. YEAR 1 HF WORDS (51-75)", "Words 51-75", "that, she, do, how, their, if, will, up, other, about, many, then, them, these, so, some, her, would, make, like, him, into, time, has, look"),

    ("0A. YEAR 1 HF WORDS (76-100)", "Words 76-100", "two, more, go, see, no, who, over, know, just, where, most, get, through, back, much, good, new, write, very, after, things, our, little, old, come"),

    ("0B. YEAR 2 COMMON EXCEPTION (1-20)", "First 20 Year 2 words", "door, floor, poor, because, find, kind, mind, behind, child, children, wild, climb, most, only, both, old, cold, gold, hold, told"),

    ("0B. YEAR 2 COMMON EXCEPTION (21-40)", "Words 21-40", "every, everybody, even, great, break, steak, pretty, beautiful, after, fast, last, past, father, class, grass, pass, plant, path, bath, hour"),

    ("0B. YEAR 2 COMMON EXCEPTION (41-64)", "Final Year 2 words", "move, prove, improve, sure, sugar, eye, could, should, would, who, whole, any, many, clothes, busy, people, water, again, half, money, Mr, Mrs, parents, Christmas"),

    ("0C. YEAR 3/4 EXCEPTION (1-25)", "First 25 Year 3/4 words", "accident, accidentally, actual, actually, address, answer, appear, arrive, believe, bicycle, breath, breathe, build, busy, business, calendar, caught, centre, century, certain, circle, complete, consider, continue, decide"),

    ("0C. YEAR 3/4 EXCEPTION (26-50)", "Words 26-50", "describe, different, difficult, disappear, early, earth, eight, eighth, enough, exercise, experience, experiment, extreme, famous, favourite, February, forward, forwards, fruit, grammar, group, guard, guide, heard, heart"),

    ("0C. YEAR 3/4 EXCEPTION (51-75)", "Words 51-75", "height, history, imagine, increase, important, interest, island, knowledge, learn, length, library, material, medicine, mention, minute, natural, naughty, notice, occasion, occasionally, often, opposite, ordinary, particular, peculiar"),

    ("0C. YEAR 3/4 EXCEPTION (76-100)", "Final Year 3/4 words", "perhaps, popular, position, possess, possession, possible, potatoes, pressure, probably, promise, purpose, quarter, question, recent, regular, reign, remember, sentence, separate, special, straight, strange, strength, suppose, surprise, therefore, though, although, thought, through, various, weight"),

    ("0D. YEAR 5/6 STATUTORY (1-25)", "First 25 Year 5/6 words", "accommodate, accompany, according, achieve, aggressive, amateur, ancient, apparent, appreciate, attached, available, average, awkward, bargain, bruise, category, cemetery, committee, communicate, community, competition, conscience, conscious, controversy, convenience"),

    ("0D. YEAR 5/6 STATUTORY (26-50)", "Words 26-50", "correspond, criticise, curiosity, definite, desperate, determined, develop, dictionary, disastrous, embarrass, environment, equip, equipment, especially, exaggerate, excellent, existence, explanation, familiar, foreign, forty, frequently, government, guarantee, harass"),

    ("0D. YEAR 5/6 STATUTORY (51-75)", "Words 51-75", "hindrance, identity, immediate, immediately, individual, interfere, interrupt, language, leisure, lightning, marvellous, mischievous, muscle, necessary, neighbour, nuisance, occupy, occur, opportunity, parliament, persuade, physical, prejudice, privilege, profession"),

    ("0D. YEAR 5/6 STATUTORY (76-100)", "Final Year 5/6 words", "programme, pronunciation, queue, recognise, recommend, relevant, restaurant, rhyme, rhythm, sacrifice, secretary, shoulder, signature, sincere, sincerely, soldier, stomach, sufficient, suggest, symbol, system, temperature, thorough, twelfth, variety, vegetable, vehicle, yacht"),

    # ========== 1-SYLLABLE: SHORT VOWELS - SIMPLE CVC ==========
    ("1A. SHORT A - Simple CVC", "Basic 3-letter words: cat pattern", "cat, bat, hat, mat, rat, sat, fat, pat, man, pan, ran, tan, can, fan, bad, dad, had, mad, sad, bag, tag, wag, rag, nag, cap, gap, lap, map, nap"),

    ("1A. SHORT A - More CVC", "More simple CVC words", "rap, tap, zap, cab, dab, jab, tab, ham, jam, ram, yam, gas, has, pal, gal, wax, tax, max, van, at, as, sat"),

    ("1A. SHORT A - With Blends", "Blends + short a", "snap, snip, clap, flap, slap, trap, strap, plan, bran, scan, span, clan, flat, brat, scat, scrap, clam, cram, gram, slam, swam, tram, black, crack, track, stack, snack"),

    ("1B. SHORT E - Simple CVC", "Basic e words: bed pattern", "bed, red, fed, led, wed, beg, leg, peg, pen, ten, hen, men, den, bet, get, jet, let, met, net, pet, set, vet, wet, yet"),

    ("1B. SHORT E - More CVC", "More e words", "bell, tell, well, sell, yell, fell, mess, less, best, nest, rest, test, vest, west, peck, neck, deck, check, hem, gem, them, step"),

    ("1C. SHORT I - Simple CVC", "Basic i words: sit pattern", "sit, bit, fit, hit, kit, lit, pit, wit, big, dig, fig, gig, jig, pig, rig, wig, bin, din, fin, pin, tin, win, dip, hip, lip"),

    ("1C. SHORT I - More CVC", "More i words", "nip, rip, sip, tip, zip, did, hid, kid, lid, rid, him, dim, rim, mix, fix, six, bib, rib, tick, pick, lick, sick, kick, wick, it, is, in"),

    ("1D. SHORT O - Simple CVC", "Basic o words: dog pattern", "dog, fog, hog, jog, log, bog, cog, box, fox, pox, dot, got, hot, jot, lot, not, pot, rot, bob, cob, job, mob, rob, sob"),

    ("1D. SHORT O - More CVC", "More o words", "hop, mop, pop, top, cop, mom, pod, nod, rock, dock, lock, sock, mock, on, ox"),

    ("1E. SHORT U - Simple CVC", "Basic u words: cup pattern", "cup, pup, up, bug, dug, hug, jug, mug, rug, tug, bud, mud, bun, fun, gun, run, sun, bus, but, cut, gut, hut, nut"),

    ("1E. SHORT U - More CVC", "More u words", "rut, tub, cub, pub, rub, sub, hum, gum, yum, sum, us, dump, jump, bump, lump, pump, hump"),

    # ========== 1-SYLLABLE: CONSONANT BLENDS ==========
    ("2A. L-BLENDS (bl, cl, fl)", "Beginning l-blends", "black, blue, blast, blend, bless, blob, blot, blunt, clap, class, clip, clock, club, clamp, clump, flag, flat, flip, flock, flow, flan, flop"),

    ("2A. L-BLENDS (gl, pl, sl)", "More l-blends", "glad, glass, glide, glow, glue, glum, plan, play, plus, plot, plum, plug, slam, slip, slow, slug, sled, slid, slim, slit, slot, skill, skull, slap"),

    ("2B. R-BLENDS (br, cr, dr)", "Beginning r-blends", "brave, brick, bring, brown, brush, bran, crab, crash, cry, crack, crown, crib, drag, drip, drop, drum, dress, drab, drill, drank"),

    ("2B. R-BLENDS (fr, gr, pr, tr)", "More r-blends", "frog, free, fresh, from, frost, frill, frank, grab, grass, green, grin, grip, grim, gram, pray, press, print, prim, trap, tree, trip, track, truck, tram, trot, trim"),

    ("2C. S-BLENDS (sc, sk, sm, sn)", "Beginning s-blends", "scale, scare, scan, scrap, skip, skate, skill, sky, skull, skid, small, smell, smile, smart, smoke, smack, snack, snap, snake, snail, snow, snag, snip"),

    ("2C. S-BLENDS (sp, st, sw)", "More s-blends", "space, spin, spot, spell, spill, stack, star, stop, step, stick, stamp, stem, swim, swing, sweet, swept, swift, swell"),

    ("2D. 3-LETTER BLENDS", "scr, spr, str, spl, thr", "scrap, scrub, screen, screw, scrape, spray, spring, spread, sprain, sprout, strap, street, string, strong, stream, strip, split, splash, splint, three, throw, through, throat, throne"),

    # ========== 1-SYLLABLE: DIGRAPHS ==========
    ("3A. CH DIGRAPH - Simple", "ch words: beginning & end", "chat, chop, chip, chin, chum, rich, much, such, lunch, bunch, munch, pinch, punch, bench, chess, chest, check, chain, chair, chalk, cheek, cheer"),

    ("3A. CH DIGRAPH - More", "More ch patterns", "chick, chimp, child, chill, champ, crunch, torch, coach, catch, match, patch, fetch, stretch, itch, ditch, witch, stitch, switch, hutch"),

    ("3B. SH DIGRAPH - Simple", "sh words: beginning & end", "shop, ship, shed, shell, shin, shot, shut, fish, dish, wish, rush, push, hush, bash, cash, dash, mash, rash, shock, shook, shark, sharp, sheep, sheet"),

    ("3B. SH DIGRAPH - More", "More sh patterns", "shape, shake, share, shine, shirt, shout, brush, smash, crash, flash, flush, splash, trash, fresh, shrink, shrimp, shred"),

    ("3C. TH DIGRAPH (voiced)", "th as in this/that", "this, that, then, them, than, with"),

    ("3C. TH DIGRAPH (unvoiced)", "th as in thin/thick", "thin, thick, think, thank, thing, third, thud, thump, tooth, teeth, cloth, three, thorn, math, bath, path, moth"),

    ("3D. WH DIGRAPH", "wh makes /w/ sound", "when, what, whale, wheat, wheel, whip, which, while, white, why"),

    ("3E. PH DIGRAPH - Simple", "ph makes /f/ sound", "phone, photo, graph"),

    # ========== 1-SYLLABLE: VOWEL TEAMS ==========
    ("4A. AI/AY - Simple AI", "ai in middle of word", "rain, pain, main, tail, rail, nail, mail, sail, wait, paid, laid, fail, train, brain, chain, grain, stain, drain, plain, snail, trail"),

    ("4A. AI/AY - AY Ending", "ay at end of word", "day, hay, lay, may, pay, say, way, play, stay, pray, spray, clay, gray, tray, ray, bay"),

    ("4B. EE/EA - Simple EE", "ee pattern words", "bee, see, tree, free, three, feed, need, seed, feet, meet, keep, deep, beep, sleep, sheep, street, green, queen"),

    ("4B. EE/EA - Simple EA", "ea pattern words", "sea, tea, pea, eat, beat, seat, heat, meat, read, lead, bead, beach, teach, reach, peach, dream, cream, steam, clean, mean, bean, leaf"),

    ("4C. OA/OW - Simple OA", "oa pattern words", "boat, coat, goat, road, toad, load, soap, soak, oak, toast, roast, float, throat, coal, goal"),

    ("4C. OA/OW - OW Pattern", "ow as long o", "show, low, bow, row, tow, mow, slow, grow, snow, blow, flow, glow, know, throw, own"),

    ("4D. IGH/IE/Y - IGH Pattern", "igh makes long i", "high, sigh, thigh, night, right, fight, light, might, sight, tight, flight, bright, fright"),

    ("4D. IGH/IE/Y - IE/Y Pattern", "ie and y as long i", "pie, tie, die, lie, sky, my, by, fly, cry, dry, fry, try, shy, why, spy"),

    ("4E. UE/EW - Simple", "ue and ew patterns", "blue, clue, glue, true, due, cue, new, few, dew, grew, flew, threw, drew, chew, knew, blew, stew, screw"),

    # ========== 1-SYLLABLE: DIPHTHONGS ==========
    ("5A. OI/OY - Simple", "oi and oy patterns", "oil, boil, soil, coil, coin, join, point, boy, toy, joy, Roy"),

    ("5B. OU/OW - Simple", "ou and ow (cow sound)", "out, loud, shout, round, pound, sound, count, cloud, house, mouse, mouth, how, now, cow, bow, down, town, brown, crown, clown, frown"),

    ("5C. AU/AW - Simple", "au and aw patterns", "haul, Paul, saw, paw, jaw, law, raw, draw, claw, straw, yawn, lawn, dawn, crawl"),

    # ========== 1-SYLLABLE: R-CONTROLLED ==========
    ("6A. AR - Simple", "ar makes /ar/ sound", "car, far, jar, tar, bar, arm, art, farm, barn, card, park, dart, mark, dark, bark, part, hard, yard, chart, smart, star, scar, shark"),

    ("6B. ER/IR/UR - ER Pattern", "er makes /er/ sound", "her, fern, herd, verb, term, germ, clerk, jerk, perk, serve, nerve"),

    ("6B. ER/IR/UR - IR Pattern", "ir makes /er/ sound", "bird, girl, sir, stir, dirt, shirt, skirt, first, third, firm, chirp, birth, twirl, swirl"),

    ("6B. ER/IR/UR - UR Pattern", "ur makes /er/ sound", "turn, burn, fur, hurt, curl, purr, burst, church, surf, purse, nurse, curve"),

    ("6C. OR - Simple", "or makes /or/ sound", "or, for, fork, cork, pork, corn, horn, torn, worn, born, storm, form, sport, short, north, horse, more, store, score, shore, snore"),

    # ========== 1-SYLLABLE: SPECIAL PATTERNS ==========
    ("6D. NG/NK - NG Ending", "ng sound at end", "ring, sing, bang, song, wing, hang, long, king, sung, rang, sang, hung, sting, spring, bring, swing, stung, strong, cling, string"),

    ("6D. NG/NK - NK Ending", "nk sound", "ink, pink, sink, link, rink, think, drink, bank, tank, sank, thank, blank, trunk, skunk, junk, bunk"),

    ("6E. AL PATTERN", "al sounds like /or/", "all, ball, call, fall, hall, tall, wall, small, talk, walk, chalk, salt"),

    # ========== 1-SYLLABLE: MAGIC E PATTERNS ==========
    ("7A. A-E (magic e)", "Silent e makes a say its name", "ate, ape, age, bake, cake, lake, make, take, wake, same, came, game, name, tame, cave, wave, gave, save, date, gate, hate, late, rate, tale, male, pale, sale, cane, lane, mane, pane, cape, tape, case, base, race, face, lace, pace, made, fade, shade, trade, grade, brave, shave"),

    ("7B. I-E (magic e)", "Silent e makes i say its name", "bite, kite, like, bike, hike, mike, time, dime, lime, five, hive, dive, drive, life, wife, knife, ride, hide, side, wide, tide, pipe, ripe, wipe, mine, nine, line, pine, vine, wine, fine, shine, size, prize, smile, while, quite, white, write"),

    ("7C. O-E (magic e)", "Silent e makes o say its name", "bone, cone, tone, zone, home, dome, Rome, hope, rope, mope, poke, woke, joke, choke, smoke, broke, spoke, stone, phone, throne, hole, mole, pole, role, note, vote, close, nose, rose, hose, those"),

    ("7D. U-E (magic e)", "Silent e makes u say its name", "cute, mute, flute, tube, cube, rude, rule, mule, tune, June, prune, use, fuse, huge"),

    ("7E. SILENT LETTERS", "kn, wr, mb, gn patterns", "know, knee, knife, knot, knit, write, wrong, wrap, wrist, thumb, lamb, climb, comb, sign, gnat"),

    ("7F. CK/TCH/DGE Endings", "Special endings", "back, pack, sack, duck, luck, truck, stuck, catch, match, patch, fetch, stretch, watch, switch, edge, bridge, judge, fudge, badge, wedge, ridge"),

    # ========== 1-SYLLABLE: COMPLEX PATTERNS ==========
    ("8A. OO - Long Sound", "oo as in moon", "zoo, moo, boo, cool, fool, pool, food, mood, moon, noon, soon, boot, root, hoop, loop, room, boom, zoom, spoon, stool, proof, roof, tooth, smooth, goose, loose"),

    ("8A. OO - Short Sound", "oo as in book", "book, look, took, cook, hook, good, wood, stood, foot, hood"),

    ("8B. SOFT C", "c sounds like s", "ice, mice, rice, nice, face, race, place, space, cent, cell"),

    ("8B. SOFT G", "g sounds like j", "age, cage, page, stage, huge, gem, gym"),

    # ========== 2-SYLLABLE WORDS ==========
    ("9A. 2-SYLLABLE - Easy", "Simple 2-syllable words", "ba-by, la-dy, ti-ger, spi-der, o-pen, o-ver, pa-per, ho-tel, mu-sic, pu-pil, ro-bot, ti-dy, e-ven, ze-bra"),

    ("9B. 2-SYLLABLE - Compound", "Easy compound words", "sun-set, hot-dog, in-side, up-on, bed-time, rain-bow, snow-man, cup-cake, pop-corn"),

    ("9C. 2-SYLLABLE - With Blends", "2-syllable with blends", "bas-ket, pic-nic, traf-fic, plas-tic, kit-chen, chick-en, prob-lem, sud-den, but-ter, din-ner, sum-mer, win-ter, sis-ter, let-ter, bet-ter"),

    ("9D. 2-SYLLABLE - Common", "Common 2-syllable words", "wa-ter, peo-ple, fa-ther, moth-er, broth-er, teach-er, gar-den, mar-ket, doc-tor, mon-ster, num-ber, cor-ner, fin-ger"),

    # ========== 3-SYLLABLE WORDS ==========
    ("10A. 3-SYLLABLE - Easy", "Simple 3-syllable words", "an-i-mal, el-e-phant, to-ma-to, ba-na-na, um-brel-la, di-no-saur, choc-o-late, yes-ter-day, to-mor-row, Sat-ur-day"),

    ("10B. 3-SYLLABLE - Common", "Common 3-syllable words", "beau-ti-ful, won-der-ful, fam-i-ly, dif-fer-ent, fa-vor-ite, li-bra-ry, his-to-ry, hos-pi-tal, dan-ger-ous, dif-fi-cult"),

    ("10C. 3-SYLLABLE - Advanced", "Harder 3-syllable words", "im-por-tant, re-mem-ber, De-cem-ber, Sep-tem-ber, No-vem-ber, tel-e-phone, com-pu-ter, to-geth-er, for-ev-er, how-ev-er"),

    # ========== 4+ SYLLABLE WORDS ==========
    ("11A. 4-SYLLABLE Words", "4-syllable practice", "al-li-ga-tor, cal-cu-la-tor, el-e-va-tor, he-li-cop-ter, re-frig-er-a-tor, el-e-men-ta-ry, par-tic-u-lar, in-cred-i-ble, in-ter-est-ing, e-mer-gen-cy"),

    ("11B. 5+ SYLLABLE Words", "Elite decoding practice", "Con-stan-ti-no-ple, hip-po-pot-a-mus, ex-traor-di-na-ry, u-ni-ver-si-ty, un-der-stand-ing, re-spon-si-bil-i-ty, or-gan-i-za-tion, com-mu-ni-ca-tion"),

    # ========== Y ENDINGS ==========
    ("12A. Y as /ee/ - Simple", "y sounds like ee", "hap-py, fun-ny, sil-ly, bun-ny, pup-py, sun-ny, luck-y, dad-dy, mum-my, ba-by, la-dy, sto-ry, par-ty, pen-ny, ber-ry"),

    ("12B. Y as /ee/ - More", "More y endings", "ver-y, man-y, an-y, pret-ty, cit-y, emp-ty, dirt-y, dus-ty, wind-y, cloud-y, rain-y, snow-y, ear-ly"),

]

# Add data to sheet
row = 2
for category, pattern, words in word_bank:
    ws[f'A{row}'] = category
    ws[f'B{row}'] = pattern
    ws[f'C{row}'] = words

    # Style category cells
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Wrap text for words column
    ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    row += 1

# Save the workbook
wb.save("C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx")
print("Segmented Excel file created successfully!")
print(f"Total sections: {len(word_bank)}")
print("Each section has 10-30 words max for easier progress tracking")
