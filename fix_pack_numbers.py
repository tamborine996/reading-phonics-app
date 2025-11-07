from openpyxl import load_workbook
import re

# Load the Excel file
wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

# Fix pack numbers (remove duplicates and add clean numbering)
pack_number = 1

for row in range(2, ws.max_row + 1):
    category_cell = ws[f'A{row}']
    current_name = category_cell.value

    if current_name:  # Only process non-empty rows
        # Remove any existing P#: prefix(es)
        clean_name = re.sub(r'^(P\d+:\s*)+', '', current_name)

        # Add clean pack number prefix
        new_name = f"P{pack_number}: {clean_name}"
        category_cell.value = new_name
        pack_number += 1

# Save the updated file
wb.save(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')

print(f"Fixed pack numbers P1 through P{pack_number - 1}")
print(f"Total packs numbered: {pack_number - 1}")
