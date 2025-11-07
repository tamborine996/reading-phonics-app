from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# Load existing workbook
input_file = r"C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx"
wb_in = load_workbook(input_file)
ws_in = wb_in.active

# Collect all words by base category (remove all the "Level" and "Part" stuff)
word_collections = defaultdict(set)  # category -> set of unique words
category_descriptions = {}  # category -> description

print("Collecting words from existing file...")

for row in range(2, ws_in.max_row + 1):
    category = ws_in[f'A{row}'].value
    pattern = ws_in[f'B{row}'].value
    words = ws_in[f'C{row}'].value

    if not category or not words:
        continue

    # Extract base category (remove "Level X", "Part X", etc.)
    base_category = category.split(' - Level')[0].split(' (Part')[0].strip()

    # Get word list
    word_list = [w.strip() for w in words.split(',') if w.strip()]

    # Add to collection (set automatically handles duplicates)
    for word in word_list:
        word_collections[base_category].add(word.lower())

    # Save description if we don't have one yet
    if base_category not in category_descriptions and pattern:
        desc = pattern.split(' - Part')[0].split(' - Level')[0].split(' - Easy')[0].split(' - Medium')[0].split(' - Hard')[0].strip()
        category_descriptions[base_category] = desc

# Create new workbook
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Word Packs"

# Headers
ws_out['A1'] = "Category"
ws_out['B1'] = "Description"
ws_out['C1'] = "Words"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
for cell in ['A1', 'B1', 'C1']:
    ws_out[cell].fill = header_fill
    ws_out[cell].font = header_font
    ws_out[cell].alignment = Alignment(horizontal='center', vertical='center')

ws_out.column_dimensions['A'].width = 40
ws_out.column_dimensions['B'].width = 40
ws_out.column_dimensions['C'].width = 90

# Process each category
row_out = 2
total_packs = 0
categories_processed = 0

print("\nCreating word packs...")

for base_category in sorted(word_collections.keys()):
    words = sorted(list(word_collections[base_category]))  # Sort alphabetically
    word_count = len(words)
    description = category_descriptions.get(base_category, "")

    print(f"\n{base_category}: {word_count} unique words")

    # Skip if too few words (less than 10)
    if word_count < 10:
        print(f"  WARNING: Skipping - too few words ({word_count})")
        continue

    categories_processed += 1

    # Split into packs of ~30 words
    pack_size = 30
    num_packs = (word_count + pack_size - 1) // pack_size  # Ceiling division

    if num_packs == 1:
        # Single pack
        ws_out[f'A{row_out}'] = base_category
        ws_out[f'B{row_out}'] = description
        ws_out[f'C{row_out}'] = ', '.join(words)

        # Style
        ws_out[f'A{row_out}'].font = Font(bold=True, size=10)
        ws_out[f'A{row_out}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_out[f'C{row_out}'].alignment = Alignment(wrap_text=True, vertical='top')

        row_out += 1
        total_packs += 1
        print(f"  Created 1 pack ({word_count} words)")
    else:
        # Multiple packs
        for pack_num in range(num_packs):
            start_idx = pack_num * pack_size
            end_idx = min(start_idx + pack_size, word_count)
            pack_words = words[start_idx:end_idx]

            ws_out[f'A{row_out}'] = f"{base_category} - Pack {pack_num + 1}"
            ws_out[f'B{row_out}'] = f"{description} (Pack {pack_num + 1} of {num_packs})"
            ws_out[f'C{row_out}'] = ', '.join(pack_words)

            # Style
            ws_out[f'A{row_out}'].font = Font(bold=True, size=10)
            ws_out[f'A{row_out}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws_out[f'C{row_out}'].alignment = Alignment(wrap_text=True, vertical='top')

            row_out += 1
            total_packs += 1

        print(f"  Created {num_packs} packs ({num_packs - 1} x {pack_size} + {len(pack_words)} words)")

# Save
output_file = r"C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx"
wb_out.save(output_file)

print("\n" + "=" * 80)
print(f"COMPLETE!")
print(f"  Categories processed: {categories_processed}")
print(f"  Total word packs created: {total_packs}")
print(f"  Pack size: ~30 words each")
print(f"  File saved: {output_file}")
print("=" * 80)
