from openpyxl import load_workbook
import json

wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

# Extract first 4 packs
packs = []

for row in range(2, 6):  # Rows 2-5 (first 4 packs)
    category = ws[f'A{row}'].value
    description = ws[f'B{row}'].value
    words = ws[f'C{row}'].value

    word_list = [w.strip() for w in words.split(',') if w.strip()]

    packs.append({
        'id': row - 1,
        'title': category,
        'description': description,
        'words': word_list
    })

# Save to JSON
with open(r'C:\Users\mqc20\Downloads\Projects\Reading app\word_packs.json', 'w') as f:
    json.dump(packs, f, indent=2)

print("Extracted first 4 packs:")
for pack in packs:
    print(f"  {pack['title']}: {len(pack['words'])} words")
