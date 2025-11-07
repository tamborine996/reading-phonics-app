from openpyxl import load_workbook

# Check current word bank
wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

all_words_current = set()
total_entries = 0

for row in range(2, ws.max_row + 1):
    words = ws[f'C{row}'].value
    if not words:
        continue

    word_list = [w.strip().lower() for w in words.split(',') if w.strip()]
    total_entries += len(word_list)
    all_words_current.update(word_list)

print(f"CURRENT WORD BANK:")
print(f"  Total word entries: {total_entries}")
print(f"  Unique words: {len(all_words_current)}")
print(f"  Total sections: {ws.max_row - 1}")
