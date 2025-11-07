from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

print("=" * 100)
print("PHONICS WORD BANK - PREVIEW")
print("=" * 100)
print()

# Show first 40 rows to give a good overview
for row in range(1, min(41, ws.max_row + 1)):
    category = ws[f'A{row}'].value or ""
    description = ws[f'B{row}'].value or ""
    words = ws[f'C{row}'].value or ""

    if row == 1:
        # Header row
        print(f"{category:45} | {description:40}")
        print("-" * 100)
    else:
        # Word count
        word_count = len([w.strip() for w in words.split(',') if w.strip()])

        # First few words preview
        word_list = [w.strip() for w in words.split(',') if w.strip()]
        preview = ', '.join(word_list[:10])
        if len(word_list) > 10:
            preview += f"... ({len(word_list)} total)"

        print(f"{category:45} | {word_count:2} words")
        print(f"  {preview}")
        print()

print("=" * 100)
print(f"Total packs: {ws.max_row - 1}")
print("=" * 100)
