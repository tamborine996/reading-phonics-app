from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Phonics Word Bank"

# Headers
ws['A1'] = "Category"
ws['B1'] = "Pattern/Description"
ws['C1'] = "Example Words"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
for cell in ['A1', 'B1', 'C1']:
    ws[cell].fill = header_fill
    ws[cell].font = header_font
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 80

# Word bank data
word_bank = [
    # HIGH FREQUENCY WORDS (100 words - Year 1)
    ("0. HIGH FREQUENCY WORDS", "100 words to know by end of Year 1", "the, that, not, look, put, and, with, then, don't, could, a, all, were, come, house, to, we, go, will, old, said, can, little, into, too, in, are, as, back, by, he, up, no, from, day, I, had, mum, children, made, of, my, one, him, time, it, her, them, Mr, I'm, was, what, do, get, if, you, there, me, just, help, they, out, down, now, Mrs, on, this, dad, came, called, she, have, big, oh, here, is, went, when, about, off, for, be, it's, got, asked, at, like, see, their, saw, his, some, looked, people, make, but, so, very, your, an"),

    # FOUNDATION - Short Vowels
    ("1. SHORT VOWEL A", "CVC words with 'a' as in cat", "cat, bat, hat, mat, rat, sat, fat, pat, van, man, pan, ran, tan, can, fan, bad, dad, had, mad, sad, bag, tag, wag, rag, nag, cap, gap, lap, map, nap, rap, tap, zap, cab, dab, jab, tab, ham, jam, ram, yam, gas, has, pal, gal, wax, tax, max"),

    ("1. SHORT VOWEL E", "CVC words with 'e' as in bed", "bed, red, fed, led, wed, beg, leg, peg, ben, den, hen, men, pen, ten, bet, get, jet, let, met, net, pet, set, vet, wet, yet, peck, neck, deck, bell, fell, tell, well, sell, yell, mess, less, best, nest, rest, test, vest, west, hem, gem, them"),

    ("1. SHORT VOWEL I", "CVC words with 'i' as in sit", "sit, bit, fit, hit, kit, lit, pit, wit, big, dig, fig, gig, jig, pig, rig, wig, bin, din, fin, pin, tin, win, dip, hip, lip, nip, rip, sip, tip, zip, did, hid, kid, lid, rid, him, dim, rim, mix, fix, six, bib, rib, tick, pick, lick, sick, kick, wick"),

    ("1. SHORT VOWEL O", "CVC words with 'o' as in dog", "dog, fog, hog, jog, log, bog, cog, box, fox, pox, dot, got, hot, jot, lot, not, pot, rot, bob, cob, job, mob, rob, sob, hop, mop, pop, top, cop, mom, pod, nod, rock, dock, lock, sock, mock, block, clock, shock"),

    ("1. SHORT VOWEL U", "CVC words with 'u' as in cup", "cup, pup, up, bug, dug, hug, jug, mug, rug, tug, bud, cud, mud, bun, fun, gun, run, sun, nun, bus, gus, pus, but, cut, gut, hut, jut, nut, rut, tub, cub, pub, rub, sub, hum, gum, yum, sum, dumb, jump, bump, lump, pump, dump"),

    # CONSONANT BLENDS - Beginning
    ("2. L-BLENDS", "bl, cl, fl, gl, pl, sl", "black, blue, blast, blend, bless, clap, class, clip, clock, club, flag, flat, flip, flock, flow, glad, glass, glide, glow, glue, plan, play, plus, plot, plum, slam, slip, slow, slug, sled"),

    ("2. R-BLENDS", "br, cr, dr, fr, gr, pr, tr", "brave, brick, bring, brown, brush, crab, crash, cry, crack, crown, drag, drip, drop, drum, dress, frog, free, fresh, from, frost, grab, grass, green, grin, grip, pray, press, print, prize, prod, trap, tree, trip, track, truck"),

    ("2. S-BLENDS", "sc, sk, sm, sn, sp, st, sw", "scale, scare, scan, skip, skate, skill, sky, skull, small, smell, smile, smart, smoke, snack, snap, snake, snail, snow, space, spin, spot, spell, spill, stack, star, stop, step, stick, swim, swing, sweet, swept"),

    ("2. 3-LETTER BLENDS", "scr, spr, str, spl, thr", "scrap, scrub, screen, screw, scrape, spray, spring, spread, sprain, sprout, strap, street, string, strong, stream, split, splash, splint, splendid, three, throw, through, throat, throne"),

    # CONSONANT DIGRAPHS
    ("3. DIGRAPH CH", "ch makes /ch/ sound", "chat, chop, chip, chin, rich, much, such, lunch, bunch, munch, chair, chain, chalk, change, check, chess, chest, child, chill, champ, church, chicken, chapter, chocolate, champion"),

    ("3. DIGRAPH SH", "sh makes /sh/ sound", "shop, ship, shed, shell, shin, dish, fish, wish, rush, push, bash, cash, dash, gash, mash, wash, shape, shake, share, shark, sharp, sheep, sheet, shelf, shine, shirt, shock, short, shout, shrink"),

    ("3. DIGRAPH TH", "th (voiced & unvoiced)", "this, that, then, them, than, math, bath, path, with, moth, cloth, thin, thick, think, thank, thing, third, thud, thumb, three, throw, throne, thunder, thoughtful"),

    ("3. DIGRAPH WH", "wh makes /w/ sound", "when, what, whack, whale, wheat, wheel, whip, which, while, whine, white, why, whisper, whistle, whisker"),

    ("3. DIGRAPH PH", "ph makes /f/ sound", "phone, photo, phase, phrase, graph, trophy, alphabet, elephant, dolphin, phantom, pharmacy, philosopher, physical, physician, geography, biography, paragraph, photographer, telephone, phonics"),

    # VOWEL TEAMS
    ("4. AI/AY (long A)", "ai in middle, ay at end", "rain, pain, main, train, brain, chain, grain, snail, trail, wait, paid, sail, tail, mail, nail, day, play, say, may, way, pay, hay, lay, ray, stay, clay, gray, pray, spray, tray"),

    ("4. EE/EA (long E)", "ee and ea patterns", "bee, see, tree, free, three, feed, need, seed, feet, meet, greet, sleep, sheep, street, green, queen,ween, beep, deep, keep, sea, tea, pea, read, lead, bead, meat, beat, seat, heat, clean, dream, stream, scream, beach, peach, teach, reach"),

    ("4. OA/OW (long O)", "oa and ow patterns", "boat, coat, goat, road, toad, load, soap, soak, oak, coach, roach, roast, toast, float, throat, show, low, bow, row, tow, mow, slow, grow, snow, blow, flow, glow, know, throw, yellow, window, elbow, shadow, pillow"),

    ("4. IGH/IE/Y (long I)", "igh, ie, y patterns", "high, sigh, thigh, night, right, fight, light, might, sight, tight, flight, bright, fright, pie, tie, die, lie, tie, sky, my, by, fly, cry, dry, fry, try, shy, why, spy, pry"),

    ("4. UE/EW (long U)", "ue and ew patterns", "blue, clue, glue, true, due, cue, rescue, value, statue, argue, continue, new, few, dew, grew, flew, threw, drew, chew, knew, blew, stew, screw, newspaper"),

    # DIPHTHONGS
    ("5. OI/OY", "oi in middle, oy at end", "oil, boil, soil, coil, coin, join, point, noise, moist, choice, voice, spoil, boy, toy, joy, Roy, soy, ploy, coy, employ, destroy, enjoy, royal, loyal, voyage"),

    ("5. OU/OW", "ou and ow (cow sound)", "out, shout, loud, cloud,ound, round, ground, pound, mound, house, mouse, mouth, south, couch, pouch, count, mount, fountain, mountain, now, how, cow, bow, wow, vow, plow, brown, crown, down, town, clown, frown, flower, power, tower, shower"),

    ("5. AU/AW", "au and aw patterns", "haul, maul, Paul, fault, vault, launch, laundry, August, author, auto, because, astronaut, saw, law, paw, raw, draw, claw, straw, thaw, yawn, lawn, dawn, fawn, crawl, brawl, shawl, awful, awkward"),

    # R-CONTROLLED VOWELS
    ("6. AR", "ar makes /ar/ sound", "car, far, jar, tar, bar, star, scar, arm, farm, harm, charm, alarm, art, cart, dart, part, start, chart, smart, park, dark, bark, mark, shark, spark, yard, card, hard, guard, garden, carpet, market, artist, partner, harvest"),

    ("6. ER/IR/UR", "all make same /er/ sound", "her, fern, herd, verb, term, germ, clerk, jerk, perk, serve, nerve, person, perfect, bird, girl, sir, stir, dirt, shirt, skirt, first, third, thirsty, circle, circus, burn, turn, hurt, fur, curl, surf, purse, nurse, turkey, turtle, surprise, purple, Thursday, furniture"),

    ("6. OR", "or makes /or/ sound", "or, for, or, fork, cork, pork, corn, horn, torn, worn, born, storm, form, sport, short, north, horse, store, more, core, score, shore, chore, snore, explorer, important, morning, corner, forget"),

    # SPECIAL PATTERNS
    ("7. SILENT E", "e makes vowel say name", "make, take, bake, cake, lake, wake, sake, brake, snake, shake, like, bike, hike, mike, spike, strike, home, dome, Rome, bone, cone, tone, stone, phone, throne, cute, mute, flute, tube, cube, rude, prune, use, fuse, refuse, confuse"),

    ("7. SILENT LETTERS", "kn, wr, mb, gn, gh", "know, knee, knife, knock, knot, knit, knight, knob, write, wrong, wrap, wrist, wreck, wrench, wring, thumb, lamb, climb, bomb, comb, crumb, numb, plumb, sign, gnat, gnaw, gnome, design, night, light, right, might, sight, flight"),

    ("7. CK/TCH/DGE", "special endings", "back, pack, sack, black, track, stick, quick, thick, trick, clock, block, shock, duck, luck, truck, stuck, catch, match, patch, latch, watch, fetch, sketch, switch, kitchen, pitcher, edge, bridge, judge, fudge, badge, wedge, lodge, ridge, budget, hedgehog"),

    # COMPLEX PATTERNS
    ("8. OO (two sounds)", "oo as in moon/book", "LONG: moon, soon, room, boom, zoom, food, mood, pool, cool, tool, school, spoon, smooth, shoot, proof, roof, goose, loose, choose, SHORT: book, look, took, cook, hook, wood, good, stood, foot, hood, shook, brook, cookie, football"),

    ("8. OUGH/AUGH", "multiple sounds!", "OUGH: through, though, thought, bought, fought, brought, cough, rough, tough, enough, dough, AUGH: caught, taught, daughter, laughter, naughty, slaughter"),

    ("8. SOFT C/G", "c=s, g=j before e,i,y", "SOFT C: cent, city, cease, cell, center, ceiling, pencil, circle, circus, cycle, bicycle, fancy, icy, mercy, SOFT G: gem, germ, giant, ginger, gym, magic, gentle, giraffe, energy, genius,atge, cage, page, stage, huge, change, charge, arrange"),

    # COMMON ENDINGS
    ("9. -ED ENDINGS", "3 sounds: /ed/, /d/, /t/", "/ED/: wanted, needed, planted, landed, ended, handed, painted, waited, /D/: played, stayed, rained, cleaned, climbed, cried, tried, smiled, /T/: jumped, helped, walked, talked, looked, cooked, picked, kicked, asked, liked"),

    ("9. -ING ENDINGS", "present progressive", "running, jumping, playing, saying, doing, going, looking, making, taking, walking, talking, reading, writing, swimming, sitting, getting, letting, stopping, shopping, hopping, skipping, spinning"),

    ("9. -S/-ES ENDINGS", "plurals & verbs", "cats, dogs, books, cars, bikes, runs, jumps, plays, boxes, glasses, wishes, catches, teaches, watches, brushes, dresses, passes, mixes, fixes, buzzes, fizzes"),

    ("9. -LE ENDINGS", "consonant + le", "able, table, cable, fable, noble, bubble, trouble, double, little, middle, riddle, fiddle, paddle, saddle, apple, purple, simple, sample, example, bottle,attle, cattle, rattle, settle, gentle, jungle, ankle, wrinkle, sprinkle, sparkle"),

    # MULTISYLLABIC WORDS - 2 Syllables
    ("10. TWO SYLLABLES", "VC/CV pattern", "napkin, basket, picnic, plastic, traffic, dentist, Contest, hundred, problem, chicken, kitchen, Signal, final, total, label, local, silent, student, music, unit, human, tulip, robot, bonus, focus, open, broken, frozen, chosen, spoken"),

    ("10. TWO SYLLABLES", "V/CV pattern", "baby, lady, gravy, navy, lazy, crazy, evil, legal, recent, hotel, motel, robot, music, human, student, silent, pilot, final, tidy, tiger, spider, virus, item, open, even, over, paper, favor, major, labor"),

    # MULTISYLLABIC WORDS - 3 Syllables
    ("11. THREE SYLLABLES", "common 3-syllable words", "animal, capital, hospital, musical, typical, magical, natural, personal, national, basketball, beautiful, butterfly, wonderful, dangerous, different, difficult, delicious, elephant,Family, powerful, popular, chocolate, Saturday, yesterday, tomorrow, favorite, library, history, mystery, energy"),

    ("11. THREE SYLLABLES", "compound patterns", "together, forever, however, whatever, remember, November, December, September, banana, pajamas, Alaska, cinema, telephone, microphone, piano, tomato, potato, volcano, Canada, Africa, America, Atlantic, computer, grandfather, grandmother, grandfather, afternoon, understand, hospital, favorite"),

    # MULTISYLLABIC WORDS - 4+ Syllables
    ("12. FOUR+ SYLLABLES", "advanced multisyllabic", "FOUR: alligator, calculator, elevator, escalator, refrigerator, helicopter, appropriate, communicate, community, electricity, information, incredible, interesting, emergency, particular, America, California, dictionary, elementary, February, January, ordinary, necessary, vocabulary, celebration, generation, operation, population, education, situation"),

    ("12. FOUR+ SYLLABLES", "elite decoding practice", "FIVE+: Constantinople, hippopotamus, congratulations, responsibility, revolutionary, extraordinary, university, uncomfortable, understanding, Massachusetts, Pennsylvania, Mississippi, refrigerator, organization, determination, participation, preparation, combination, multiplication, investigation, imagination, communication, entertainment, encyclopedia, photographer, absolutely, unfortunately, enthusiastically"),

    # ADVANCED VOCABULARY
    ("13. ADVANCED WORDS", "complex real-world words", "photosynthesis, metamorphosis, democracy, geography, biography, autobiography, telescope, microscope, telephone, symphony, pharmacy, philosophy, catastrophe, apostrophe, atmosphere, hemisphere, thermometer, chronometer, architecture, agriculture, manufacture, literature, temperature, adventure, creature, feature, furniture, measure, treasure, pleasure, picture, nature, capture"),

    ("13. SCIENTIFIC TERMS", "science vocabulary", "experiment, hypothesis, laboratory, microscope, telescope, molecule, nucleus, chromosome, ecosystem, environment, biodiversity, photosynthesis, respiration, evaporation, precipitation, constellation, dinosaur, archaeology, paleontology, geology, meteorology, astronomy, biology, chemistry, physics, mathematics, calculation, equation, fraction, multiplication, division"),
]

# Add data to sheet
row = 2
for category, pattern, words in word_bank:
    ws[f'A{row}'] = category
    ws[f'B{row}'] = pattern
    ws[f'C{row}'] = words

    # Style category cells
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Wrap text for words column
    ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    row += 1

# Save the workbook
wb.save("C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx")
print("Excel file created successfully: Phonics_Word_Bank.xlsx")
