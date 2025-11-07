from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

# Add sequential pack numbers (P1, P2, P3, etc.)
pack_number = 1

for row in range(2, ws.max_row + 1):
    category_cell = ws[f'A{row}']
    current_name = category_cell.value

    if current_name:  # Only process non-empty rows
        # Add pack number prefix
        new_name = f"P{pack_number}: {current_name}"
        category_cell.value = new_name
        pack_number += 1

# Save the updated file
wb.save(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')

print(f"Added sequential pack numbers P1 through P{pack_number - 1}")
print(f"Total packs numbered: {pack_number - 1}")
