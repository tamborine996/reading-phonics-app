from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import re

def count_syllables(word):
    """Rough syllable counter"""
    word = word.lower().strip()
    vowels = "aeiouy"
    syllable_count = 0
    previous_was_vowel = False

    for i, char in enumerate(word):
        is_vowel = char in vowels
        if is_vowel and not previous_was_vowel:
            syllable_count += 1
        previous_was_vowel = is_vowel

    # Adjust for silent e
    if word.endswith('e') and syllable_count > 1:
        syllable_count -= 1

    return max(1, syllable_count)

def get_difficulty_score(word):
    """Calculate difficulty: syllables, length, complexity"""
    syllables = count_syllables(word)
    length = len(word)

    # Base score on syllables (most important)
    score = syllables * 10

    # Add points for length
    if length > 10:
        score += 5
    elif length > 7:
        score += 3
    elif length > 5:
        score += 1

    return score

def split_by_difficulty(words_list, category_base, pattern_base):
    """Split words into difficulty levels"""
    if not words_list:
        return []

    # Calculate difficulty for each word
    word_scores = [(word, get_difficulty_score(word)) for word in words_list]
    word_scores.sort(key=lambda x: x[1])  # Sort by difficulty

    # Determine difficulty breaks
    scores = [s for w, s in word_scores]
    min_score = min(scores)
    max_score = max(scores)

    if max_score - min_score <= 5:
        # All similar difficulty - just split by quantity
        results = []
        chunks = [word_scores[i:i+35] for i in range(0, len(word_scores), 35)]

        if len(chunks) == 1:
            return [(f"{category_base}", pattern_base,
                    ', '.join([w for w, s in chunks[0]]))]
        else:
            for idx, chunk in enumerate(chunks):
                letter = chr(65 + idx)  # A, B, C...
                results.append((
                    f"{category_base} - Level 1{letter}",
                    f"{pattern_base} - Part {idx+1} (same difficulty)",
                    ', '.join([w for w, s in chunk])
                ))
            return results

    # Multiple difficulty levels
    results = []

    # Level 1: Easiest (1 syllable, simple)
    level1 = [w for w, s in word_scores if s <= 15]
    # Level 2: Medium (2 syllables or complex 1-syllable)
    level2 = [w for w, s in word_scores if 15 < s <= 25]
    # Level 3: Hard (3+ syllables or very complex)
    level3 = [w for w, s in word_scores if s > 25]

    # Process Level 1
    if level1:
        chunks = [level1[i:i+35] for i in range(0, len(level1), 35)]
        if len(chunks) == 1:
            results.append((f"{category_base} - Level 1",
                          f"{pattern_base} - Easy (1 syllable)",
                          ', '.join(chunks[0])))
        else:
            for idx, chunk in enumerate(chunks):
                letter = chr(65 + idx)
                results.append((f"{category_base} - Level 1{letter}",
                              f"{pattern_base} - Easy Part {idx+1}",
                              ', '.join(chunk)))

    # Process Level 2
    if level2:
        chunks = [level2[i:i+35] for i in range(0, len(level2), 35)]
        if len(chunks) == 1:
            results.append((f"{category_base} - Level 2",
                          f"{pattern_base} - Medium (2 syllables)",
                          ', '.join(chunks[0])))
        else:
            for idx, chunk in enumerate(chunks):
                letter = chr(65 + idx)
                results.append((f"{category_base} - Level 2{letter}",
                              f"{pattern_base} - Medium Part {idx+1}",
                              ', '.join(chunk)))

    # Process Level 3
    if level3:
        chunks = [level3[i:i+35] for i in range(0, len(level3), 35)]
        if len(chunks) == 1:
            results.append((f"{category_base} - Level 3",
                          f"{pattern_base} - Hard (3+ syllables)",
                          ', '.join(chunks[0])))
        else:
            for idx, chunk in enumerate(chunks):
                letter = chr(65 + idx)
                results.append((f"{category_base} - Level 3{letter}",
                              f"{pattern_base} - Hard Part {idx+1}",
                              ', '.join(chunk)))

    return results

# Load existing workbook
input_file = "C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx"
wb_in = load_workbook(input_file)
ws_in = wb_in.active

# Track all words to find duplicates
all_words = defaultdict(list)  # word -> list of categories it appears in
unique_words_by_category = {}  # category -> set of unique words

# First pass: collect all words and find duplicates
print("Analyzing for duplicates...")
for row in range(2, ws_in.max_row + 1):
    category = ws_in[f'A{row}'].value
    words = ws_in[f'C{row}'].value

    if not category or not words:
        continue

    word_list = [w.strip().lower() for w in words.split(',') if w.strip()]

    for word in word_list:
        all_words[word].append(category)

# Find duplicates
duplicates = {word: cats for word, cats in all_words.items() if len(cats) > 1}
print(f"Found {len(duplicates)} duplicate words!")
print("\nSample duplicates:")
for i, (word, cats) in enumerate(list(duplicates.items())[:10]):
    print(f"  '{word}' appears in: {', '.join(cats[:3])}...")

# Create new workbook
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Organized Word Bank"

# Headers
ws_out['A1'] = "Category & Level"
ws_out['B1'] = "Pattern/Difficulty"
ws_out['C1'] = "Words"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
for cell in ['A1', 'B1', 'C1']:
    ws_out[cell].fill = header_fill
    ws_out[cell].font = header_font
    ws_out[cell].alignment = Alignment(horizontal='center', vertical='center')

ws_out.column_dimensions['A'].width = 45
ws_out.column_dimensions['B'].width = 45
ws_out.column_dimensions['C'].width = 90

# Reorganize data
row_out = 2
processed_words = set()  # Track to avoid duplicates

print("\nReorganizing with difficulty levels...")

for row_in in range(2, ws_in.max_row + 1):
    category = ws_in[f'A{row_in}'].value
    pattern = ws_in[f'B{row_in}'].value
    words = ws_in[f'C{row_in}'].value

    if not category or not words:
        continue

    # Skip if this is already a split section (from previous split)
    if "(Part " in category:
        # Extract base category
        category = category.split(" (Part")[0]

    # Get unique words for this category (remove duplicates within category)
    word_list = [w.strip() for w in words.split(',') if w.strip()]
    unique_words = []
    for word in word_list:
        word_lower = word.lower()
        # Keep word if it's the first time we see it, or if it's in a high-freq/exception list
        if word_lower not in processed_words or "FREQUENCY" in category or "EXCEPTION" in category or "STATUTORY" in category:
            unique_words.append(word)
            processed_words.add(word_lower)

    if not unique_words:
        continue

    # Split by difficulty
    sections = split_by_difficulty(unique_words, category, pattern)

    for cat, pat, word_str in sections:
        ws_out[f'A{row_out}'] = cat
        ws_out[f'B{row_out}'] = pat
        ws_out[f'C{row_out}'] = word_str

        # Style
        ws_out[f'A{row_out}'].font = Font(bold=True, size=10)
        ws_out[f'A{row_out}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_out[f'C{row_out}'].alignment = Alignment(wrap_text=True, vertical='top')

        row_out += 1

# Save
output_file = "C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx"
wb_out.save(output_file)

print(f"\nCompleted!")
print(f"Total sections: {row_out - 2}")
print(f"Words organized by difficulty levels:")
print(f"  - Level 1A/1B/1C = Easy (same difficulty, split for size)")
print(f"  - Level 2 = Medium difficulty")
print(f"  - Level 3 = Hard difficulty")
print(f"Duplicates removed (kept first occurrence)")
print(f"File: {output_file}")

# Create duplicate report
with open("C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Duplicate_Report.txt", "w") as f:
    f.write("DUPLICATE WORDS REPORT\n")
    f.write("=" * 50 + "\n\n")
    f.write(f"Total duplicate words found: {len(duplicates)}\n\n")

    for word, cats in sorted(duplicates.items()):
        f.write(f"'{word}' appears in:\n")
        for cat in cats:
            f.write(f"  - {cat}\n")
        f.write("\n")

print("\nDuplicate report saved to: Duplicate_Report.txt")
