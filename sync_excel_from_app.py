"""
Sync Excel from app.js
This script reads the wordPacks data from app.js and updates the Excel file to match.
Excel becomes a mirror of what's actually live on the website.
"""

import re
import json
from openpyxl import load_workbook

# Read app.js and extract wordPacks data
print("Reading app.js...")
with open(r'C:\Users\mqc20\Downloads\Projects\Reading app\app.js', 'r', encoding='utf-8') as f:
    content = f.read()

# Extract the wordPacks array using regex
# Looking for: let wordPacks = [ ... ];
match = re.search(r'let wordPacks = (\[[\s\S]*?\]);', content)
if not match:
    print("ERROR: Could not find wordPacks array in app.js")
    exit(1)

# Get the JSON array (it's valid JSON inside the JavaScript)
wordpacks_json = match.group(1)

# Parse it
try:
    word_packs = json.loads(wordpacks_json)
    print(f"Found {len(word_packs)} packs in app.js")
except json.JSONDecodeError as e:
    print(f"ERROR parsing wordPacks JSON: {e}")
    exit(1)

# Load Excel file
print("Loading Excel file...")
wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

# Update Excel rows to match app.js data
print("Syncing Excel with app.js data...")
for pack in word_packs:
    pack_id = pack['id']
    title = pack['title']
    description = pack['description']
    words = pack['words']

    # Excel row is pack_id + 1 (because row 1 is header)
    row_num = pack_id + 1

    # Update the row
    ws[f'A{row_num}'] = title
    ws[f'B{row_num}'] = description
    ws[f'C{row_num}'] = ', '.join(words)

    print(f"  Updated P{pack_id}: {title} ({len(words)} words)")

# Save Excel file
print("Saving Excel file...")
wb.save(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')

print("\nSUCCESS! Excel now mirrors what's live in app.js")
print(f"Updated {len(word_packs)} packs in Excel")
