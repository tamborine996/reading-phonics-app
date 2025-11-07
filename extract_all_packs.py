"""
Extract ALL 130 packs from Excel
Creates a complete JSON file with all packs organized by sub-packs
"""

from openpyxl import load_workbook
import json
import re

# Load Excel
wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

print(f"Reading Excel file with {ws.max_row - 1} packs...")

# Extract all packs
all_packs = []
pack_number = 1

for row in range(2, ws.max_row + 1):
    category = ws[f'A{row}'].value
    description = ws[f'B{row}'].value
    words = ws[f'C{row}'].value

    if not category or not words:
        continue

    word_list = [w.strip() for w in words.split(',') if w.strip()]

    # Remove the P# prefix from category to get clean name
    clean_category = re.sub(r'^P\d+:\s*', '', category)

    pack = {
        'id': pack_number,
        'title': f"P{pack_number}: {clean_category}",
        'description': description if description else f"{len(word_list)} words",
        'category': clean_category,
        'words': word_list
    }

    all_packs.append(pack)
    pack_number += 1

    if pack_number <= 5:  # Show first few
        print(f"  P{pack['id']}: {clean_category} ({len(word_list)} words)")

print(f"\nExtracted {len(all_packs)} packs")
print(f"Total words: {sum(len(p['words']) for p in all_packs)}")

# Save to JSON for reference
with open(r'C:\Users\mqc20\Downloads\Projects\Reading app\all_packs_extracted.json', 'w', encoding='utf-8') as f:
    json.dump(all_packs, f, indent=2, ensure_ascii=False)

print(f"\nSaved to all_packs_extracted.json")
