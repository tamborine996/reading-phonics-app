from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load the existing comprehensive workbook
input_file = "C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx"
wb_in = load_workbook(input_file)
ws_in = wb_in.active

# Create new workbook for split version
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Word Bank (Split)"

# Copy headers
ws_out['A1'] = "Category"
ws_out['B1'] = "Pattern/Description"
ws_out['C1'] = "Words (max 40 per section)"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
for cell in ['A1', 'B1', 'C1']:
    ws_out[cell].fill = header_fill
    ws_out[cell].font = header_font
    ws_out[cell].alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
ws_out.column_dimensions['A'].width = 40
ws_out.column_dimensions['B'].width = 42
ws_out.column_dimensions['C'].width = 90

def split_words(word_string, max_words=40):
    """Split a comma-separated word string into chunks"""
    words = [w.strip() for w in word_string.split(',')]
    chunks = []
    for i in range(0, len(words), max_words):
        chunk = words[i:i + max_words]
        chunks.append(', '.join(chunk))
    return chunks

# Read all data from comprehensive file (skip header row)
row_out = 2
for row_in in range(2, ws_in.max_row + 1):
    category = ws_in[f'A{row_in}'].value
    pattern = ws_in[f'B{row_in}'].value
    words = ws_in[f'C{row_in}'].value

    if not category or not words:
        continue

    # Count words
    word_list = [w.strip() for w in words.split(',')]
    word_count = len(word_list)

    if word_count <= 40:
        # Small enough, keep as one section
        ws_out[f'A{row_out}'] = category
        ws_out[f'B{row_out}'] = pattern
        ws_out[f'C{row_out}'] = words

        # Style
        ws_out[f'A{row_out}'].font = Font(bold=True, size=10)
        ws_out[f'A{row_out}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_out[f'C{row_out}'].alignment = Alignment(wrap_text=True, vertical='top')

        row_out += 1
    else:
        # Split into multiple parts
        word_chunks = split_words(words, max_words=40)

        for idx, chunk in enumerate(word_chunks, 1):
            new_category = f"{category} (Part {idx}/{len(word_chunks)})"
            new_pattern = f"{pattern} - Part {idx} of {len(word_chunks)}"

            ws_out[f'A{row_out}'] = new_category
            ws_out[f'B{row_out}'] = new_pattern
            ws_out[f'C{row_out}'] = chunk

            # Style
            ws_out[f'A{row_out}'].font = Font(bold=True, size=10)
            ws_out[f'A{row_out}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws_out[f'C{row_out}'].alignment = Alignment(wrap_text=True, vertical='top')

            row_out += 1

# Save
output_file = "C:\\Users\\mqc20\\Downloads\\Projects\\Reading app\\Phonics_Word_Bank.xlsx"
wb_out.save(output_file)

print(f"Split word bank created!")
print(f"Total sections: {row_out - 2}")
print(f"Large sections split into max 40 words each")
print(f"ALL words preserved, NO hyphens added")
print(f"File: {output_file}")
