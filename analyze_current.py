from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\mqc20\Downloads\Projects\Reading app\Phonics_Word_Bank.xlsx')
ws = wb.active

print("CURRENT WORD BANK ANALYSIS")
print("=" * 100)
print(f"{'Category':<50} {'Word Count':>10}")
print("-" * 100)

for i in range(2, ws.max_row + 1):
    category = ws[f'A{i}'].value
    pattern = ws[f'B{i}'].value
    words = ws[f'C{i}'].value

    if not category or not words:
        continue

    word_count = len([w.strip() for w in words.split(',') if w.strip()])
    print(f"{category:<50} {word_count:>10}")

print("\n" + "=" * 100)
print(f"Total sections: {ws.max_row - 1}")
